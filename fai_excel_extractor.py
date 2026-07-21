#!/usr/bin/env python3
"""Extract IPQC/FAI sampling rows from inspection Excel workbooks.

The reader follows the same idea as the 2177PQE scripts: it reads workbook XML
directly so the result does not depend on Excel formula cache recalculation.
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from xml.etree import ElementTree as ET
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NS_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

OUTPUT_BASE_COLUMNS = ["Date ", "Sampling process", "Sampling time", "Sampling line#/Machine#", "Furnace No.", "FAI"]
CPK_REQUIREMENT = 1.33
METADATA_PREFIX = "__"


def col_to_num(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + ord(ch.upper()) - 64
    return n


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def numeric_value(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def normalize_label(value: Any) -> str:
    return clean_text(value).replace(" ", "").replace("\u3000", "").replace("：", ":")


def parse_cell_range(ref: str) -> Tuple[int, int, int, int]:
    if ":" not in ref:
        match = re.match(r"([A-Z]+)(\d+)$", ref)
        if not match:
            raise ValueError("Bad cell reference: %s" % ref)
        col = col_to_num(match.group(1))
        row = int(match.group(2))
        return row, col, row, col
    start, end = ref.split(":", 1)
    s_match = re.match(r"([A-Z]+)(\d+)$", start)
    e_match = re.match(r"([A-Z]+)(\d+)$", end)
    if not s_match or not e_match:
        raise ValueError("Bad cell range: %s" % ref)
    return int(s_match.group(2)), col_to_num(s_match.group(1)), int(e_match.group(2)), col_to_num(e_match.group(1))


class ExcelXmlReader:
    """Small OOXML reader for values, sheet names and merged-cell ranges."""

    def __init__(self, path: Path):
        self.path = path
        self.zip = ZipFile(path)
        self.shared_strings = self._read_shared_strings()
        self.sheet_targets = self._read_sheet_targets()

    def close(self) -> None:
        self.zip.close()

    def __enter__(self) -> "ExcelXmlReader":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def _read_shared_strings(self) -> List[str]:
        if "xl/sharedStrings.xml" not in self.zip.namelist():
            return []
        root = ET.fromstring(self.zip.read("xl/sharedStrings.xml"))
        strings: List[str] = []
        for si in root.findall(NS_MAIN + "si"):
            strings.append("".join(t.text or "" for t in si.iter(NS_MAIN + "t")))
        return strings

    def _read_sheet_targets(self) -> Dict[str, str]:
        workbook = ET.fromstring(self.zip.read("xl/workbook.xml"))
        rels = ET.fromstring(self.zip.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        targets: Dict[str, str] = {}
        sheets = workbook.find(NS_MAIN + "sheets")
        if sheets is None:
            return targets
        for sheet in sheets:
            name = sheet.attrib["name"]
            rid = sheet.attrib[NS_REL + "id"]
            target = rid_to_target[rid]
            if not target.startswith("xl/"):
                target = "xl/" + target.lstrip("/")
            targets[name] = target
        return targets

    def sheet_names(self) -> List[str]:
        return list(self.sheet_targets.keys())

    def cell_value(self, cell: ET.Element) -> Any:
        cell_type = cell.attrib.get("t")
        if cell_type == "inlineStr":
            return "".join(t.text or "" for t in cell.iter(NS_MAIN + "t"))
        if cell_type == "str":
            v = cell.find(NS_MAIN + "v")
            return "" if v is None else (v.text or "")
        v = cell.find(NS_MAIN + "v")
        if v is None:
            return ""
        text = v.text or ""
        if cell_type == "s":
            try:
                return self.shared_strings[int(text)]
            except (ValueError, IndexError):
                return text
        if cell_type == "b":
            return text == "1"
        try:
            return float(text) if any(ch in text for ch in ".Ee") else int(text)
        except ValueError:
            return text

    def read_sheet(self, sheet_name: str) -> Tuple[Dict[int, Dict[int, Any]], List[Tuple[int, int, int, int]]]:
        target = self.sheet_targets[sheet_name]
        root = ET.fromstring(self.zip.read(target))
        rows: Dict[int, Dict[int, Any]] = {}
        sheet_data = root.find(NS_MAIN + "sheetData")
        if sheet_data is not None:
            for row in sheet_data.findall(NS_MAIN + "row"):
                row_num = int(row.attrib.get("r", "0"))
                values: Dict[int, Any] = {}
                for cell in row.findall(NS_MAIN + "c"):
                    ref = cell.attrib.get("r", "")
                    match = re.match(r"([A-Z]+)", ref)
                    if not match:
                        continue
                    values[col_to_num(match.group(1))] = self.cell_value(cell)
                if any(clean_text(v) != "" for v in values.values()):
                    rows[row_num] = values
        merges: List[Tuple[int, int, int, int]] = []
        merge_cells = root.find(NS_MAIN + "mergeCells")
        if merge_cells is not None:
            for merge in merge_cells.findall(NS_MAIN + "mergeCell"):
                ref = merge.attrib.get("ref")
                if ref:
                    try:
                        merges.append(parse_cell_range(ref))
                    except ValueError:
                        pass
        return rows, merges


@dataclass(frozen=True)
class SamplingGroup:
    columns: Tuple[int, ...]
    sample_date: date
    sample_time: time


def merged_lookup(merges: Sequence[Tuple[int, int, int, int]], rows_of_interest: Iterable[int]) -> Dict[Tuple[int, int], Tuple[int, int]]:
    wanted = set(rows_of_interest)
    lookup: Dict[Tuple[int, int], Tuple[int, int]] = {}
    for min_row, min_col, max_row, max_col in merges:
        hit_rows = [r for r in wanted if min_row <= r <= max_row]
        for row in hit_rows:
            for col in range(min_col, max_col + 1):
                lookup[(row, col)] = (min_row, min_col)
    return lookup


def value_at(rows: Dict[int, Dict[int, Any]], row: int, col: int, merge_map: Dict[Tuple[int, int], Tuple[int, int]]) -> Any:
    if row in rows and col in rows[row]:
        direct_value = rows[row][col]
        if clean_text(direct_value) != "":
            return direct_value
    top_left = merge_map.get((row, col))
    if top_left is None:
        return rows.get(row, {}).get(col, "")
    top_row, top_col = top_left
    return rows.get(top_row, {}).get(top_col, "")


def parse_yymmdd(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        text = "%06d" % int(value)
    else:
        text = re.sub(r"\D", "", clean_text(value))
        if len(text) == 8 and text.startswith("20"):
            return date(int(text[:4]), int(text[4:6]), int(text[6:8]))
    if len(text) != 6:
        return None
    yy, mm, dd = int(text[:2]), int(text[2:4]), int(text[4:6])
    year = 2000 + yy if yy < 70 else 1900 + yy
    try:
        return date(year, mm, dd)
    except ValueError:
        return None


def parse_hhmm(value: Any) -> Optional[time]:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, timedelta):
        total_seconds = int(value.total_seconds()) % (24 * 3600)
        return time(total_seconds // 3600, (total_seconds % 3600) // 60)
    if isinstance(value, (int, float)):
        if value < 0:
            return None
        total_seconds = int(round((float(value) % 1) * 24 * 3600))
        return time((total_seconds // 3600) % 24, (total_seconds % 3600) // 60)
    text = clean_text(value).replace("：", ":")
    match = re.search(r"(\d{1,2})\s*:\s*(\d{1,2})", text)
    if not match:
        return None
    hour, minute = int(match.group(1)), int(match.group(2))
    if hour == 24:
        hour = 0
    if 0 <= hour <= 23 and 0 <= minute <= 59:
        return time(hour, minute)
    return None


def excel_time_fraction(t: time) -> float:
    return (t.hour * 3600 + t.minute * 60 + t.second) / 86400.0


def display_fai(value: Any, prefix_fai: bool = False) -> Any:
    text = clean_text(value)
    if prefix_fai and text and not text.upper().startswith("FAI"):
        return "FAI" + text
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def contains_chinese(value: Any) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", clean_text(value)))


def is_rejected_sample(value: Any) -> bool:
    text = clean_text(value)
    normalized = text.replace(" ", "").replace("\u3000", "").upper()
    return normalized in ("OK", "NG", "/", "\\", "-", "--", "NA", "N/A", "NULL", "NONE") or contains_chinese(value)


def is_meaningful_sample(value: Any) -> bool:
    """Return True only for real measurement data.

    Template result rows such as appearance OK/NG rows, judgment rows and empty
    placeholder rows should not be exported as Sample data.
    """
    text = clean_text(value)
    if text == "":
        return False
    normalized = text.replace(" ", "").replace("\u3000", "").upper()
    if is_rejected_sample(value):
        return False
    return True


def find_row_containing(rows: Dict[int, Dict[int, Any]], keywords: Sequence[str]) -> Optional[int]:
    for row_num in sorted(rows):
        for value in rows[row_num].values():
            label = normalize_label(value)
            if all(keyword in label for keyword in keywords):
                return row_num
    return None


def find_label_row(rows: Dict[int, Dict[int, Any]], labels: Sequence[str]) -> Optional[int]:
    for row_num in sorted(rows):
        for value in rows[row_num].values():
            label = normalize_label(value)
            if any(expected in label for expected in labels):
                return row_num
    return None


def find_header_value(rows: Dict[int, Dict[int, Any]], merge_map: Dict[Tuple[int, int], Tuple[int, int]], labels: Sequence[str], max_row: int = 8) -> str:
    max_col = max((col for row in rows.values() for col in row.keys()), default=0)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if any(label in normalize_label(value_at(rows, row, col, merge_map)) for label in labels):
                for next_col in range(col + 1, min(max_col, col + 4) + 1):
                    value = clean_text(value_at(rows, row, next_col, merge_map))
                    if value:
                        return value
    return ""


def find_dimension_header(rows: Dict[int, Dict[int, Any]]) -> Optional[Tuple[int, int]]:
    for row_num in sorted(rows):
        for col, value in rows[row_num].items():
            if "尺寸序号" in normalize_label(value):
                return row_num, col
    return None


def find_spec_columns(
    rows: Dict[int, Dict[int, Any]],
    header_row: int,
    merge_map: Dict[Tuple[int, int], Tuple[int, int]],
) -> Dict[str, int]:
    max_col = max((col for row in rows.values() for col in row.keys()), default=0)
    labels = {
        "Nominal": ("标准值",),
        "Upper tolerance": ("上公差",),
        "Lower tolerance": ("下公差",),
        "LSL": ("LSL",),
        "USL": ("USL",),
        "Limit LSL": ("LIMITLSL", "LIMIT_LSL"),
        "Limit USL": ("LIMITUSL", "LIMIT_USL"),
    }
    spec_cols: Dict[str, int] = {}
    for row in (header_row, header_row + 1):
        for col in range(1, max_col + 1):
            label = normalize_label(value_at(rows, row, col, merge_map)).upper()
            if not label:
                continue
            for key, expected_labels in labels.items():
                if key not in spec_cols and any(expected in label for expected in expected_labels):
                    spec_cols[key] = col
    return spec_cols


def build_sampling_groups(
    rows: Dict[int, Dict[int, Any]],
    date_row: int,
    time_row: int,
    merge_map: Dict[Tuple[int, int], Tuple[int, int]],
) -> List[SamplingGroup]:
    max_col = max((col for row in rows.values() for col in row.keys()), default=0)
    groups: List[SamplingGroup] = []
    current_cols: List[int] = []
    current_key: Optional[Tuple[date, time]] = None
    for col in range(1, max_col + 1):
        sample_date = parse_yymmdd(value_at(rows, date_row, col, merge_map))
        sample_time = parse_hhmm(value_at(rows, time_row, col, merge_map))
        key = (sample_date, sample_time) if sample_date is not None and sample_time is not None else None
        if key is not None and key == current_key:
            current_cols.append(col)
            continue
        if current_key is not None and current_cols:
            groups.append(SamplingGroup(tuple(current_cols), current_key[0], current_key[1]))
        current_cols = [col] if key is not None else []
        current_key = key
    if current_key is not None and current_cols:
        groups.append(SamplingGroup(tuple(current_cols), current_key[0], current_key[1]))
    return groups


def extract_sheet_records(sheet_name: str, rows: Dict[int, Dict[int, Any]], merges: Sequence[Tuple[int, int, int, int]], prefix_fai: bool = False) -> List[Dict[str, Any]]:
    dim_header = find_dimension_header(rows)
    date_row = find_row_containing(rows, ["检验", "日期"]) or find_label_row(rows, ["出炉日期", "日期(yymmdd)", "日期"])
    time_row = find_row_containing(rows, ["检验", "时间"]) or find_label_row(rows, ["出炉时间", "时间(hh:mm)", "时间"])
    furnace_row = find_label_row(rows, ["炉号"])
    if dim_header is None or date_row is None or time_row is None:
        return []

    header_row, fai_col = dim_header
    rows_of_interest = set([4, date_row, time_row])
    if furnace_row is not None:
        rows_of_interest.add(furnace_row)
    rows_of_interest.update(r for r in rows.keys() if r >= header_row)
    merge_map = merged_lookup(merges, rows_of_interest)
    groups = build_sampling_groups(rows, date_row, time_row, merge_map)
    if not groups:
        return []

    spec_cols = find_spec_columns(rows, header_row, merge_map)

    process = ""
    if "工序" in normalize_label(value_at(rows, 4, 7, merge_map)):
        process = clean_text(value_at(rows, 4, 8, merge_map))
    if not process:
        process = find_header_value(rows, merge_map, ["工序"])
    machine = ""
    if "机台" in normalize_label(value_at(rows, 4, 9, merge_map)):
        machine = clean_text(value_at(rows, 4, 10, merge_map))
    if not machine:
        machine = find_header_value(rows, merge_map, ["机台", "机台号"])

    records: List[Dict[str, Any]] = []
    for row_num in sorted(r for r in rows.keys() if r > header_row):
        raw_fai = value_at(rows, row_num, fai_col, merge_map)
        if clean_text(raw_fai) == "":
            continue
        if normalize_label(raw_fai) in ("尺寸序号", "判定", "检验员"):
            continue
        for group in groups:
            raw_samples = [value_at(rows, row_num, col, merge_map) for col in group.columns]
            filled_samples = [sample for sample in raw_samples if clean_text(sample) != ""]
            if not filled_samples or any(is_rejected_sample(sample) for sample in filled_samples):
                continue
            samples = [sample for sample in filled_samples if is_meaningful_sample(sample)]
            if not samples:
                continue
            record: Dict[str, Any] = {
                "source_sheet": sheet_name,
                "Date ": group.sample_date,
                "Sampling process": process,
                "Sampling time": excel_time_fraction(group.sample_time),
                "Sampling line#/Machine#": machine,
                "Furnace No.": clean_text(value_at(rows, furnace_row, group.columns[0], merge_map)) if furnace_row is not None else "",
                "FAI": display_fai(raw_fai, prefix_fai=prefix_fai),
            }
            for spec_name, spec_col in spec_cols.items():
                spec_value = numeric_value(value_at(rows, row_num, spec_col, merge_map))
                if spec_value is not None:
                    record[METADATA_PREFIX + spec_name] = spec_value
            for idx, sample in enumerate(samples, start=1):
                record["Sample %d" % idx] = sample
            records.append(record)
    return records


def extract_workbook(path: Path, prefix_fai: bool = False) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    records: List[Dict[str, Any]] = []
    sheet_counts: Dict[str, int] = {}
    with ExcelXmlReader(path) as reader:
        for sheet_name in reader.sheet_names():
            rows, merges = reader.read_sheet(sheet_name)
            sheet_records = extract_sheet_records(sheet_name, rows, merges, prefix_fai=prefix_fai)
            if sheet_records:
                for record in sheet_records:
                    record["source_file"] = path.name
                records.extend(sheet_records)
                sheet_counts[sheet_name] = len(sheet_records)
    return records, sheet_counts


def extract_workbooks(paths: Sequence[Path], prefix_fai: bool = False) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    records: List[Dict[str, Any]] = []
    counts: Dict[str, int] = {}
    for path in paths:
        workbook_records, workbook_counts = extract_workbook(path, prefix_fai=prefix_fai)
        records.extend(workbook_records)
        for sheet_name, count in workbook_counts.items():
            key = "%s / %s" % (path.name, sheet_name)
            counts[key] = count
    return records, counts


def diagnose_workbook(path: Path) -> List[str]:
    """Return short diagnostics for templates that produce no measurement rows."""
    messages: List[str] = []
    with ExcelXmlReader(path) as reader:
        sheet_names = reader.sheet_names()
        messages.append("Sheets: " + ", ".join(sheet_names))
        for sheet_name in sheet_names:
            rows, _ = reader.read_sheet(sheet_name)
            dim_header = find_dimension_header(rows)
            date_row = find_row_containing(rows, ["检验", "日期"]) or find_label_row(rows, ["出炉日期", "日期(yymmdd)", "日期"])
            time_row = find_row_containing(rows, ["检验", "时间"]) or find_label_row(rows, ["出炉时间", "时间(hh:mm)", "时间"])
            if dim_header or date_row or time_row:
                messages.append(
                    "%s: dimension_header=%s, date_row=%s, time_row=%s"
                    % (sheet_name, dim_header, date_row, time_row)
                )
    return messages


def sample_values(record: Dict[str, Any]) -> List[float]:
    values: List[float] = []
    for key, value in record.items():
        if not key.startswith("Sample "):
            continue
        number = numeric_value(value)
        if number is not None:
            values.append(number)
    return values


def sample_stdev(values: Sequence[float]) -> Optional[float]:
    if len(values) < 2:
        return None
    avg = sum(values) / len(values)
    variance = sum((value - avg) ** 2 for value in values) / (len(values) - 1)
    return math.sqrt(variance)


def round_number(value: Optional[float], digits: int = 6) -> Any:
    if value is None:
        return ""
    if math.isinf(value):
        return "NA"
    return round(value, digits)


def calculate_cpk(mean: float, stdev: Optional[float], lsl: Optional[float], usl: Optional[float]) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if stdev is None:
        return None, None, None
    if stdev == 0:
        ucpk = math.inf if usl is None or mean <= usl else -math.inf
        lcpk = math.inf if lsl is None or mean >= lsl else -math.inf
    else:
        ucpk = None if usl is None else (usl - mean) / (3 * stdev)
        lcpk = None if lsl is None else (mean - lsl) / (3 * stdev)
    if ucpk is None and lcpk is None:
        cpk = None
    elif ucpk is None:
        cpk = lcpk
    elif lcpk is None:
        cpk = ucpk
    else:
        cpk = min(ucpk, lcpk)
    return ucpk, lcpk, cpk


def first_numeric(records: Sequence[Dict[str, Any]], key: str) -> Optional[float]:
    for record in records:
        number = numeric_value(record.get(key))
        if number is not None:
            return number
    return None


def build_cpk_summary(records: Sequence[Dict[str, Any]], requirement: float = CPK_REQUIREMENT) -> List[Dict[str, Any]]:
    groups: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = {}
    for record in records:
        if not sample_values(record):
            continue
        key = (
            clean_text(record.get("source_file")),
            clean_text(record.get("source_sheet")),
            clean_text(record.get("Sampling process")),
            clean_text(record.get("Sampling line#/Machine#")),
            clean_text(record.get("FAI")),
        )
        groups.setdefault(key, []).append(record)

    summary: List[Dict[str, Any]] = []
    for key in sorted(groups.keys(), key=lambda item: [natural_part(part) for part in item]):
        group_records = groups[key]
        values = [value for record in group_records for value in sample_values(record)]
        if len(values) < 2:
            continue
        mean = sum(values) / len(values)
        stdev = sample_stdev(values)
        nominal = first_numeric(group_records, METADATA_PREFIX + "Nominal")
        lsl = first_numeric(group_records, METADATA_PREFIX + "Limit LSL")
        usl = first_numeric(group_records, METADATA_PREFIX + "Limit USL")
        if lsl is None:
            lsl = first_numeric(group_records, METADATA_PREFIX + "LSL")
        if usl is None:
            usl = first_numeric(group_records, METADATA_PREFIX + "USL")
        if nominal is None and lsl is not None and usl is not None:
            nominal = (lsl + usl) / 2
        ucpk, lcpk, cpk = calculate_cpk(mean, stdev, lsl, usl)
        cpk_failed = cpk is not None and not math.isinf(cpk) and cpk < requirement

        proposed_tol: Optional[float] = None
        proposed_lsl: Optional[float] = None
        proposed_usl: Optional[float] = None
        proposed_ucpk: Optional[float] = None
        proposed_lcpk: Optional[float] = None
        proposed_cpk: Optional[float] = None
        if cpk_failed and nominal is not None and stdev is not None:
            proposed_tol = abs(mean - nominal) if stdev == 0 else requirement * 3 * stdev + abs(mean - nominal)
            proposed_lsl = nominal - proposed_tol
            proposed_usl = nominal + proposed_tol
            proposed_ucpk, proposed_lcpk, proposed_cpk = calculate_cpk(mean, stdev, proposed_lsl, proposed_usl)

        source_file, source_sheet, process, machine, fai = key
        summary.append(
            {
                "Source file": source_file,
                "Source sheet": source_sheet,
                "Sampling process": process,
                "Sampling line#/Machine#": machine,
                "FAI": fai,
                "Nominal": round_number(nominal),
                "LSL": round_number(lsl),
                "USL": round_number(usl),
                "Count": len(values),
                "Average": round_number(mean),
                "StdDev": round_number(stdev),
                "Min": round_number(min(values)),
                "Max": round_number(max(values)),
                "UCPK": round_number(ucpk, 3),
                "LCPK": round_number(lcpk, 3),
                "CPK": round_number(cpk, 3),
                "CPK requirement": requirement,
                "Status": "NG" if cpk_failed else "OK",
                "Proposed Lower Tol": round_number(-proposed_tol if proposed_tol is not None else None),
                "Proposed Upper Tol": round_number(proposed_tol),
                "Proposed LSL": round_number(proposed_lsl),
                "Proposed USL": round_number(proposed_usl),
                "Recalc UCPK": round_number(proposed_ucpk, 3),
                "Recalc LCPK": round_number(proposed_lcpk, 3),
                "Recalc CPK": round_number(proposed_cpk, 3),
            }
        )
    return summary


def natural_part(value: Any) -> List[Any]:
    text = clean_text(value)
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", text)]


def write_cpk_summary_sheet(wb: Workbook, records: Sequence[Dict[str, Any]]) -> None:
    summary = build_cpk_summary(records)
    if not summary:
        return
    ws = wb.create_sheet("CPK summary")
    columns = list(summary[0].keys())
    header_fill = PatternFill("solid", fgColor="FCE4D6")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")
    center = Alignment(horizontal="center", vertical="center")
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for row_idx, row in enumerate(summary, start=2):
        row_fill = fail_fill if row.get("Status") == "NG" else ok_fill
        for col_idx, header in enumerate(columns, start=1):
            cell = ws.cell(row_idx, col_idx, row.get(header, ""))
            cell.font = body_font
            cell.alignment = center
            if header == "Status":
                cell.fill = row_fill
    for col_idx, header in enumerate(columns, start=1):
        values = [clean_text(ws.cell(row, col_idx).value) for row in range(1, min(ws.max_row, 200) + 1)]
        width = max([len(header)] + [len(v) for v in values if v]) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 26)
    ws.freeze_panes = "A2"


def export_workbook(records: Sequence[Dict[str, Any]], output_path: Path, include_source_sheet: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "IPQC template"

    max_samples = 0
    for record in records:
        for key in record.keys():
            match = re.match(r"Sample (\d+)$", key)
            if match:
                max_samples = max(max_samples, int(match.group(1)))
    columns = list(OUTPUT_BASE_COLUMNS)
    columns.extend("Sample %d" % idx for idx in range(1, max(10, max_samples) + 1))
    source_files = {clean_text(record.get("source_file")) for record in records if clean_text(record.get("source_file"))}
    include_source_file = include_source_sheet or len(source_files) > 1
    if include_source_file:
        columns.append("Source file")
    if include_source_sheet:
        columns.append("Source sheet")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_idx, record in enumerate(records, start=2):
        for col_idx, header in enumerate(columns, start=1):
            key = "source_sheet" if header == "Source sheet" else "source_file" if header == "Source file" else header
            cell = ws.cell(row_idx, col_idx, record.get(key, ""))
            cell.font = body_font
            cell.alignment = center
            if header == "Date ":
                cell.number_format = "yyyy-mm-dd"
            elif header == "Sampling time":
                cell.number_format = "hh:mm"

    for col_idx, header in enumerate(columns, start=1):
        values = [clean_text(ws.cell(row, col_idx).value) for row in range(1, min(ws.max_row, 200) + 1)]
        width = max([len(header)] + [len(v) for v in values if v]) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 32)
    ws.freeze_panes = "A2"
    write_cpk_summary_sheet(wb, records)
    wb.save(output_path)


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(input_path.stem + "_extracted.xlsx")


def default_multi_output_path(input_paths: Sequence[Path]) -> Path:
    if len(input_paths) == 1:
        return default_output_path(input_paths[0])
    base_dir = input_paths[0].parent if input_paths else Path.cwd()
    return base_dir / "FAI_IPQC_Excel_Extracted.xlsx"


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract all inspection sheet FAI sampling data to a new Excel file.")
    parser.add_argument("input", type=Path, nargs="+", help="One or more input .xlsx/.xlsm inspection workbooks")
    parser.add_argument("-o", "--output", type=Path, help="Output .xlsx path")
    parser.add_argument("--prefix-fai", action="store_true", help="Prefix FAI values with 'FAI' (for example 4 -> FAI4)")
    parser.add_argument("--include-source-sheet", action="store_true", help="Append a Source sheet column for traceability")
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    input_paths = [path.expanduser().resolve() for path in args.input]
    output_path = (args.output or default_multi_output_path(input_paths)).expanduser().resolve()
    for input_path in input_paths:
        if not input_path.exists():
            print("Input file does not exist: %s" % input_path, file=sys.stderr)
            return 2
    try:
        records, sheet_counts = extract_workbooks(input_paths, prefix_fai=args.prefix_fai)
    except BadZipFile as exc:
        print("Input is not a valid .xlsx/.xlsm workbook: %s" % exc, file=sys.stderr)
        return 2
    if not records:
        print("No matching inspection records found.", file=sys.stderr)
        return 1
    export_workbook(records, output_path, include_source_sheet=args.include_source_sheet)
    print("Output: %s" % output_path)
    print("Records: %d" % len(records))
    for sheet_name, count in sheet_counts.items():
        print("  %s: %d" % (sheet_name, count))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())